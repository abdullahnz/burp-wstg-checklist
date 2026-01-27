#!/usr/bin/python3 

from openpyxl import load_workbook
import json

wstg_file = "WSTG-Checklist_v4.2.xlsx"

def parse_excel_checklist():
    wb = load_workbook(wstg_file)
    
    # Select sheet (active sheet)
    sheet = wb.active
    checklist_title = ""

    data = []

    # Read all rows
    for row in sheet.iter_rows(values_only=True):
        if not row[0]:
            continue

        if row[1] == "Test Name":
            checklist_title = row[0]

        if not row[0].startswith("WSTG"):
            continue

        checklist_id = row[0]
        checklist_name = row[1]
        checklist_objectives = row[2] or ""

        checklist_objectives = checklist_objectives.replace("- ", "")

        data.append({
            "title": checklist_title,
            "id": checklist_id,
            "name": checklist_name,
            "objectives": checklist_objectives.split("\n"),
        })


    # with open("wstg-checklist-v4.2.json", "w") as f:
    #     f.write(json.dumps(data))

    return data


def main():
    parsed = parse_excel_checklist()

    data = {}
    for row in parsed:
        # if not data.get(row['title']):
        #     data[row['title']] = [] 

        # data[row['title']].append({
        #     'id': row['id'],
        #     'name': row['name'],
        # })
        print(f'Checklist("{row['id']}", "{row['title']}", "{row['name']}"),')

    
    # with open("data.json", "w") as f:
    #     f.write(json.dumps(data))



if __name__ == '__main__':
    main()    